#!/usr/bin/env python3
"""
Court Filing Analyzer
Analyzes court filings to extract key information and outputs to Excel.
"""

import os
import sys
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional

# Document processing
import docx
import pdfplumber
from PyPDF2 import PdfReader

# LLM integration
from openai import OpenAI

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class CourtFilingAnalyzer:
    """Analyzes court filings and extracts structured information."""
    
    def __init__(self, output_dir: str = "output"):
        """Initialize the analyzer."""
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.client = OpenAI()  # API key pre-configured in environment
        
        # Theme colors for Excel
        self.THEME = {
            'primary': '2D2D2D',
            'light': 'E5E5E5',
            'accent': '1F4E79',
            'header_bg': '2D2D2D',
            'header_text': 'FFFFFF'
        }
        
    def extract_text_from_file(self, file_path: str) -> str:
        """Extract text content from various document formats."""
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        extension = file_path.suffix.lower()
        
        try:
            if extension == '.docx':
                return self._extract_from_docx(file_path)
            elif extension == '.pdf':
                return self._extract_from_pdf(file_path)
            elif extension == '.txt':
                return self._extract_from_txt(file_path)
            else:
                raise ValueError(f"Unsupported file format: {extension}")
        except Exception as e:
            print(f"Error extracting text from {file_path}: {e}")
            raise
    
    def _extract_from_docx(self, file_path: Path) -> str:
        """Extract text from DOCX file."""
        doc = docx.Document(file_path)
        return '\n'.join([paragraph.text for paragraph in doc.paragraphs])
    
    def _extract_from_pdf(self, file_path: Path) -> str:
        """Extract text from PDF file."""
        text_parts = []
        
        # Try pdfplumber first (better text extraction)
        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        text_parts.append(text)
            return '\n'.join(text_parts)
        except Exception:
            # Fallback to PyPDF2
            reader = PdfReader(file_path)
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    text_parts.append(text)
            return '\n'.join(text_parts)
    
    def _extract_from_txt(self, file_path: Path) -> str:
        """Extract text from TXT file."""
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    
    def analyze_filing(self, text: str, filename: str) -> Dict[str, Any]:
        """Analyze court filing text using LLM to extract structured information."""
        
        prompt = f"""You are a legal document analyzer. Analyze the following court filing and extract structured information.

Document text:
{text}

Extract the following information in JSON format. If a field cannot be determined, use null for text fields, or "N/A" for classification fields.

Required fields:
- case_number: The case number (string)
- case_name: The name of the case (string)
- court: The court where the case was filed (string)
- judges: Judge(s) presiding over the case (string, comma-separated if multiple)
- docket_id: Unique identifier for this docket entry (string)
- date_filed: When the document was filed (string in YYYY-MM-DD format, or null)
- document_type: Type of filing (string)
- filing_party: Name of the party filing the document (string)
- title: Name/title of the filing (string)
- description: Brief summary of the filing (string, 1-2 sentences)
- key_procedural_event: Does the filing impact case progression? (string: "Yes" or "No")
- procedural_events_referenced: Date and description of key procedural events (string)
- date_issued: If this is a court order, date issued (string in YYYY-MM-DD format, or null)
- order_type: If this is a court order, type of order (string or null)
- impact_on_case: If this is a court order, how it affects case progression (string or null)
- multidistrict_litigation: Does it relate to an MDL? (string: "Yes", "No", or "N/A")

Return ONLY valid JSON with these exact field names, no additional text."""

        try:
            response = self.client.chat.completions.create(
                model="gpt-4.1-mini",
                messages=[
                    {"role": "system", "content": "You are a legal document analyzer that extracts structured information from court filings. Always respond with valid JSON only."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=2000
            )
            
            result_text = response.choices[0].message.content.strip()
            
            # Remove markdown code blocks if present
            if result_text.startswith('```'):
                result_text = result_text.split('```')[1]
                if result_text.startswith('json'):
                    result_text = result_text[4:]
                result_text = result_text.strip()
            
            result = json.loads(result_text)
            result['source_file'] = filename
            return result
            
        except json.JSONDecodeError as e:
            print(f"Error parsing LLM response as JSON: {e}")
            print(f"Response was: {result_text}")
            raise
        except Exception as e:
            print(f"Error analyzing filing: {e}")
            raise
    
    def create_excel_report(self, results: List[Dict[str, Any]], output_filename: str = None):
        """Create a professional Excel report with the analysis results."""
        
        if not results:
            print("No results to export.")
            return
        
        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"court_filing_analysis_{timestamp}.xlsx"
        
        output_path = self.output_dir / output_filename
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Court Filing Analysis"
        
        # Hide gridlines
        ws.sheet_view.showGridLines = False
        
        # Set up left margin
        ws.column_dimensions['A'].width = 3
        
        # Title
        ws['B2'] = "COURT FILING ANALYSIS"
        ws['B2'].font = Font(name='Calibri', size=18, bold=True, color=self.THEME['primary'])
        
        # Metadata
        ws['B3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['B3'].font = Font(name='Calibri', size=10, color='666666')
        
        ws['B4'] = f"Total Filings Analyzed: {len(results)}"
        ws['B4'].font = Font(name='Calibri', size=10, color='666666')
        
        # Define columns
        columns = [
            ("Source File", 25),
            ("Case Number", 18),
            ("Case Name", 35),
            ("Court", 30),
            ("Judge(s)", 25),
            ("Docket ID", 15),
            ("Date Filed", 12),
            ("Document Type", 20),
            ("Filing Party", 25),
            ("Title", 35),
            ("Description", 50),
            ("Key Procedural Event", 15),
            ("Procedural Events", 40),
            ("Date Issued", 12),
            ("Order Type", 20),
            ("Impact on Case", 40),
            ("MDL", 10)
        ]
        
        # Header row
        header_row = 6
        for col_idx, (col_name, col_width) in enumerate(columns, start=2):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = Font(name='Calibri', size=11, bold=True, color=self.THEME['header_text'])
            cell.fill = PatternFill(start_color=self.THEME['header_bg'], end_color=self.THEME['header_bg'], fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = Border(
                bottom=Side(style='thin', color='FFFFFF')
            )
            
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = col_width
        
        # Data rows
        field_mapping = [
            'source_file',
            'case_number',
            'case_name',
            'court',
            'judges',
            'docket_id',
            'date_filed',
            'document_type',
            'filing_party',
            'title',
            'description',
            'key_procedural_event',
            'procedural_events_referenced',
            'date_issued',
            'order_type',
            'impact_on_case',
            'multidistrict_litigation'
        ]
        
        for row_idx, result in enumerate(results, start=header_row + 1):
            for col_idx, field in enumerate(field_mapping, start=2):
                value = result.get(field)
                
                # Handle None/null values
                if value is None or value == 'null':
                    value = ''
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = Font(name='Calibri', size=10)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                
                # Conditional formatting for classification fields
                if field == 'key_procedural_event':
                    if value == 'Yes':
                        cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                        cell.font = Font(name='Calibri', size=10, color='155724', bold=True)
                    elif value == 'No':
                        cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
                        cell.font = Font(name='Calibri', size=10, color='721C24')
                
                if field == 'multidistrict_litigation':
                    if value == 'Yes':
                        cell.fill = PatternFill(start_color='D1ECF1', end_color='D1ECF1', fill_type='solid')
                        cell.font = Font(name='Calibri', size=10, color='0C5460', bold=True)
                
                # Add borders
                cell.border = Border(
                    bottom=Side(style='thin', color='E5E5E5')
                )
        
        # Freeze panes (freeze header row)
        ws.freeze_panes = ws['B7']
        
        # Auto-adjust row heights
        for row in range(header_row + 1, header_row + len(results) + 1):
            ws.row_dimensions[row].height = 40
        
        # Save workbook
        wb.save(output_path)
        print(f"\nExcel report saved to: {output_path}")
        return output_path
    
    def process_files(self, file_paths: List[str]) -> str:
        """Process multiple court filing files and generate Excel report."""
        
        results = []
        
        print(f"\nProcessing {len(file_paths)} file(s)...\n")
        
        for idx, file_path in enumerate(file_paths, 1):
            file_path = Path(file_path)
            print(f"[{idx}/{len(file_paths)}] Processing: {file_path.name}")
            
            try:
                # Extract text
                text = self.extract_text_from_file(file_path)
                
                if not text.strip():
                    print(f"  ⚠ Warning: No text extracted from {file_path.name}")
                    continue
                
                # Analyze with LLM
                result = self.analyze_filing(text, file_path.name)
                results.append(result)
                print(f"  ✓ Analysis complete")
                
            except Exception as e:
                print(f"  ✗ Error: {e}")
                continue
        
        if not results:
            print("\n✗ No files were successfully processed.")
            return None
        
        # Generate Excel report
        print(f"\nGenerating Excel report...")
        output_path = self.create_excel_report(results)
        
        print(f"\n{'='*60}")
        print(f"✓ Analysis complete!")
        print(f"  Files processed: {len(results)}/{len(file_paths)}")
        print(f"  Output: {output_path}")
        print(f"{'='*60}\n")
        
        return str(output_path)


def main():
    """Main entry point for the program."""
    
    if len(sys.argv) < 2:
        print("Usage: python3 court_analyzer.py <file1> [file2] [file3] ...")
        print("\nSupported formats: .pdf, .docx, .txt")
        print("\nExample:")
        print("  python3 court_analyzer.py input/filing1.pdf input/filing2.docx")
        sys.exit(1)
    
    file_paths = sys.argv[1:]
    
    # Validate files exist
    for file_path in file_paths:
        if not Path(file_path).exists():
            print(f"Error: File not found: {file_path}")
            sys.exit(1)
    
    # Run analysis
    analyzer = CourtFilingAnalyzer()
    analyzer.process_files(file_paths)


if __name__ == "__main__":
    main()
